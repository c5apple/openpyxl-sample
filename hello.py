# coding:utf-8

from openpyxl import Workbook

# ワークブック取得
wb = Workbook()

# アクティブワークシート取得
ws = wb.active

# 値を入力する
ws['A1'] = 42
ws['B1'] = 'こんにちは'

# まとめて最終行に入力する
ws.append([1, 2, 3])

# 日付
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
